"""AUDITORIA SOLO LECTURA: el Excel de la contadora contra el modulo de Planilla.

🔴 POR QUE EXISTE. La quincena del 16 al 30 de julio de 2026 (`2026-07-2`) es la
ULTIMA que la contadora calculo A MANO, antes de que el modulo existiera. Es la
unica vara que hay. Este script pone las dos columnas al lado y dice en que se
diferencian, renglon por renglon, para que Daniel decida que se cambia.

🔑 SE LEEN LAS FORMULAS, NO SOLO LOS VALORES (`data_only=False` ademas de
`data_only=True`). La formula es la que dice la REGLA; el valor solo dice el
resultado. Sin eso no se ve, por ejemplo, que a RODRIGO MIRANDA los seguros se
le escribieron a mano (17.06) en vez de salir del `*9.75%` como a todos los
demas, ni que la columna «Exedente de 9 horas» esta en cero en las TRES
empresas porque la contadora no la usa.

── COMO SE CORRE ────────────────────────────────────────────────────────────
    # 1) el lado del modulo, por la puerta de la app (no un SELECT crudo):
    DOTENV_CONFIG_PATH=.env.local npx tsx -r dotenv/config \
        scripts/_diag-planilla-vs-contadora.ts 2026-07-2 > /tmp/modulo.txt
    # 2) el cruce:
    python3 scripts/_diag-excel-contadora-vs-planilla.py /tmp/modulo.txt

Los Excel se buscan solos en ~/Downloads. Nada de esto escribe en ningun lado.
"""
import os
import re
import sys
import glob

try:
    import openpyxl
except ImportError:
    sys.exit("falta openpyxl: pip3 install openpyxl")

DESCARGAS = os.path.expanduser("~/Downloads")

# Los tres archivos que mando la contadora. Los nombres traen espacios dobles y
# un «VIST ANA» partido, asi que se buscan por patron y no por nombre exacto.
ARCHIVOS = [
    ("confecciones_boston", "BOSTON*Planilla Quincenal*julio*2026.xlsx"),
    ("vistana", "VIST*ANA*Planilla Quincenal*julio*2026.xlsx"),
    ("fashion_wear", "FASHION WEAR*Planilla Quincenal*julio*2026.xlsx"),
]

# El Excel escribe el nombre a mano en cada fila; el modulo lo trae de la ficha.
# 🩸 No se puede juntar por texto: «LUZ LOPEZ» en el Excel es «LUZ BOSQUEZ» en la
# ficha, y «Julio Guzman» de Vistana es el mismo «Julio Garay» que cobra aparte
# en Fashion Wear. El puente se escribe a mano y se revisa a mano.
ALIAS = {
    "BRICEIDA MONTERO": "8", "ALEJANDRA CAMANO": "22", "KENNY VARGAS": "28",
    "KENER HERNANDEZ": "17", "DOMINGO HENRIQUEZ": "32", "CRISTIAN BLANCO": "25",
    "JULICAR CORONA": "15", "ANDRES A. GONZALEZ C.": "23", "RAMON MIRANDA": "21",
    "LUZ LOPEZ": "18", "LAURA L. CASIANO V.": "38", "SAMIR POLO ARRIETA": "42",
    "MARTA A. CHAVARRIA": "43", "CARLOS RUIZ": "44", "LUIS BALLESTA": "46",
    "MARIA BETHANCOURT": "49", "HECTOR L. PEREZ A": "48",
    "YERITZA Y. SOLIS CASTRO": "51", "JENNIFER ARMAS": None,
    "JULIO GUZMAN": "11", "LUCIA ANGELA GARCIA": "7", "EDWIN GOMEZ": None,
    "ROXANA HERNANDEZ": "1", "RODRIGO MIRANDA": "13", "LUIS ADRIAN ARROYO": "9",
    "ANDREA PEREZ": "16", "JORMAN HERNANDEZ": "5",
    "ELOYN MENDOZA": "29", "LUIS PARAJON": "10", "ESMER CRUZ": "12",
    "KEVIN LUBO": "6", "CARLOS BALTODANO": "37", "JHONY FLORES": "40",
    "SAMUEL GOMEZ": "41", "JULIO GARAY": "11",
}

# Las columnas del cuadro grande, iguales en las tres empresas.
COLS = {"D": "qnal", "E": "ex125", "F": "ausencias", "G": "tardanzas",
        "H": "ex150", "I": "excedente", "J": "domingos", "K": "feriados",
        "L": "bruto", "M": "segsoc", "N": "segedu", "O": "isr",
        "P": "prestamo", "Q": "terceros", "R": "mercancia",
        "S": "totded", "T": "otros", "U": "neto"}


def sin_tildes(s):
    tabla = str.maketrans("ÁÉÍÓÚÜÑáéíóúüñ", "AEIOUUNaeiouun")
    return re.sub(r"\s+", " ", str(s).translate(tabla)).strip().upper()


def num(v):
    return float(v) if isinstance(v, (int, float)) else 0.0


def leer_excel(ruta):
    """Devuelve las filas del/los cuadro(s) de una empresa.

    Un archivo puede traer DOS cuadros: «Planilla» y «Servicios Profesionales».
    Se detectan por el encabezado «Nombre | Cargo | Salario Mensual», no por
    numero de fila: Boston los tiene en 5, Vistana en 6 y 21, y Fashion Wear
    deja el primero vacio con un titulo de 2024 sin borrar.
    """
    wf = openpyxl.load_workbook(ruta, data_only=False)
    wv = openpyxl.load_workbook(ruta, data_only=True)
    hoja = next((h for h in wf.sheetnames if "julio" in h.lower()), wf.sheetnames[0])
    sf, sv = wf[hoja], wv[hoja]
    filas, encabezados = [], []
    for r in range(1, sf.max_row + 1):
        if sin_tildes(sf.cell(r, 1).value or "") == "NOMBRE":
            encabezados.append(r)
    for i, hr in enumerate(encabezados):
        fin = encabezados[i + 1] if i + 1 < len(encabezados) else sf.max_row + 1
        for r in range(hr + 1, fin):
            nombre = sf.cell(r, 1).value
            if not nombre or sin_tildes(nombre) in ("", "TOTALES"):
                continue
            if sin_tildes(sf.cell(r, 2).value or "") == "TOTALES":
                continue
            fila = {"nombre": sin_tildes(nombre), "fila": r, "hoja": hoja,
                    "bloque": "serv_prof" if i else "planilla",
                    "salario_mensual": num(sv.cell(r, 3).value), "amano": [], "formulas": {}}
            # 🩸 NO se descarta la fila que da todo cero. ELOYN MENDOZA sale en
            # cero con la nota «Vacaciones» y esa fila ES el dato: la contadora
            # decidio no pagarle, no se olvido de el.
            for letra, campo in COLS.items():
                c = openpyxl.utils.column_index_from_string(letra)
                f, v = sf.cell(r, c).value, num(sv.cell(r, c).value)
                fila[campo] = v
                fila["formulas"][campo] = f
                # 🔴 EL HALLAZGO CENTRAL: monto escrito a mano donde el resto de
                # la columna lleva formula. Es un parche, y hay que listarlo.
                if v and not (isinstance(f, str) and f.startswith("=")):
                    fila["amano"].append(f"{campo}={v}")
            if fila["salario_mensual"] or any(fila[c] for c in COLS.values()):
                filas.append(fila)
    return filas


def leer_modulo(ruta):
    """Lee la salida de `_diag-planilla-vs-contadora.ts`. Nunca la base."""
    por_codigo, empresa = {}, None
    for ln in open(ruta, encoding="utf-8"):
        m = re.match(r"EMPRESA: (\S+)", ln)
        if m:
            empresa = m.group(1)
            continue
        m = re.match(r"(.{26}) (\S+)\s+SIN DINERO -> (.*)", ln)
        if m:
            por_codigo[m.group(2)] = {"nombre": m.group(1).strip(), "empresa": empresa,
                                      "sin_dinero": m.group(3).strip()}
            continue
        m = re.match(r"(.{26}) (\S{1,4})\s+((?:\s*-?[\d.]+){19})\s*$", ln)
        if m:
            n = [float(x) for x in m.group(3).split()]
            campos = ["qnal", "ex125", "ausencias", "tardanzas", "ex150", "excedente",
                      "domingos", "feriados", "bruto", "segsoc", "segedu", "isr",
                      "prestamo", "terceros", "mercancia", "totded", "otros", "neto"]
            d = dict(zip(campos, n[:9] + n[9:]))
            d.update(nombre=m.group(1).strip(), empresa=empresa, sin_dinero=None)
            por_codigo[m.group(2)] = d
    return por_codigo


def main():
    ruta_mod = sys.argv[1] if len(sys.argv) > 1 else "/tmp/modulo.txt"
    if not os.path.exists(ruta_mod):
        sys.exit(f"no encuentro la salida del modulo en {ruta_mod} (ver el encabezado)")
    mod = leer_modulo(ruta_mod)

    gran_total = {"excel": 0.0, "modulo": 0.0, "dif_bruto": 0.0, "abs_bruto": 0.0}
    for empresa, patron in ARCHIVOS:
        cands = glob.glob(os.path.join(DESCARGAS, patron))
        if not cands:
            print(f"\n⚠️  no encontre el Excel de {empresa} ({patron})")
            continue
        filas = leer_excel(cands[0])
        print("\n" + "=" * 118)
        print(f"EMPRESA {empresa}   ({os.path.basename(cands[0])})")
        print("=" * 118)
        cab = ["PERSONA", "COD", "BLOQUE", "qnal-X", "qnal-M", "bruto-X", "bruto-M",
               "Δbruto", "neto-X", "neto-M", "Δneto"]
        print(f"{cab[0]:<26}{cab[1]:>4} {cab[2]:<10}" + "".join(f"{c:>10}" for c in cab[3:]))
        tx = tm = 0.0
        comp = {"n": 0, "dif": 0.0, "abs": 0.0}
        vistos = set()
        for f in filas:
            cod = ALIAS.get(f["nombre"], "?")
            m = mod.get(cod) if cod else None
            if cod:
                vistos.add(cod)
            # 🔴 SOLO CUENTA SI EL MODULO LO PUSO EN ESTA MISMA EMPRESA. JULIO
            # cobra en Vistana Y en Fashion Wear en el Excel; el modulo lo tiene
            # entero en Vistana. Sumarle su linea a las dos empresas inventaria
            # un pago que no existe y taparia justo el hallazgo.
            mismo = bool(m) and m.get("empresa") == empresa and not m.get("sin_dinero")
            dm = m if mismo else {k: 0.0 for k in COLS.values()}
            tx += f["neto"]
            tm += dm["neto"] if mismo else 0.0
            # 🔴 EL HUECO QUE MIDE UNA REGLA DEL CALCULO ES EL DEL **BRUTO**, y
            # solo sobre los renglones COMPARABLES. El neto arrastra prestamos,
            # terceros y mercancia que la contadora teclea a mano y el modulo no
            # tiene cargados: mirarlo ahi hace que un arreglo del calculo
            # aparezca como si empeorara. Se cuentan la suma con signo (para que
            # se vea si el modulo paga de mas o de menos) y la suma de valores
            # absolutos (para que un +30 y un -30 no se tapen entre si).
            if mismo:
                comp["n"] += 1
                comp["dif"] += dm["bruto"] - f["bruto"]
                comp["abs"] += abs(dm["bruto"] - f["bruto"])
            nota = ""
            if not cod:
                nota = "  ← NO EXISTE EN EL MODULO"
            elif m is None:
                nota = "  ← el modulo no lo listo en ninguna empresa"
            elif m.get("sin_dinero"):
                nota = f"  ← modulo SIN DINERO: {m['sin_dinero'][:60]}"
            elif not mismo:
                nota = f"  ← el modulo lo paga entero en {m['empresa']} (neto {m['neto']:.2f})"
            print(f"{f['nombre'][:26]:<26}{cod or '-':>4} {f['bloque']:<10}"
                  + "".join(f"{v:>10.2f}" for v in [
                      f["qnal"], dm["qnal"], f["bruto"], dm["bruto"],
                      dm["bruto"] - f["bruto"], f["neto"], dm["neto"],
                      dm["neto"] - f["neto"]]) + nota)
            if f["amano"]:
                print(f"{'':>31}🔴 escrito A MANO (sin formula): " + " · ".join(f["amano"]))
        # Quien esta en el modulo y NO en el Excel.
        for cod, m in mod.items():
            if m.get("empresa") == empresa and cod not in vistos:
                est = m.get("sin_dinero") or f"neto {m['neto']:.2f}"
                print(f"{m['nombre'][:26]:<26}{cod:>4} {'—':<10}"
                      f"{'':>80}  ← EN EL MODULO Y NO EN EL EXCEL ({est})")
        print(f"{'TOTAL NETO':<26}{'':>4} {'':<10}{'':>60}{tx:>10.2f}{tm:>10.2f}{tm - tx:>10.2f}")
        print(f"{'HUECO DEL BRUTO':<26}{'':>4} {'':<10}"
              f"solo los {comp['n']} renglones comparables:"
              f"   con signo {comp['dif']:+.2f}   en valor absoluto {comp['abs']:.2f}")
        gran_total["excel"] += tx
        gran_total["modulo"] += tm
        gran_total["dif_bruto"] += comp["dif"]
        gran_total["abs_bruto"] += comp["abs"]

    d = gran_total["modulo"] - gran_total["excel"]
    print("\n" + "=" * 118)
    print(f"GRAN TOTAL NETO   Excel {gran_total['excel']:.2f}   modulo {gran_total['modulo']:.2f}"
          f"   diferencia {d:+.2f}")
    print(f"GRAN TOTAL BRUTO (comparables)   con signo {gran_total['dif_bruto']:+.2f}"
          f"   en valor absoluto {gran_total['abs_bruto']:.2f}")
    reglas_del_excel()


def reglas_del_excel():
    """Las REGLAS que solo se ven leyendo las formulas, no los valores.

    Cada una es algo que el Excel hace y el modulo no (o al reves). Se imprimen
    con el monto al lado: un hallazgo sin monto no sirve para decidir.
    """
    print("\n" + "=" * 118)
    print("REGLAS QUE SALEN DE LAS FORMULAS")
    print("=" * 118)
    for empresa, patron in ARCHIVOS:
        cands = glob.glob(os.path.join(DESCARGAS, patron))
        if not cands:
            continue
        wf = openpyxl.load_workbook(cands[0], data_only=False)
        wv = openpyxl.load_workbook(cands[0], data_only=True)

        # ── 1. La columna «Exedente de 9 horas» (el 2,625) ─────────────────
        # 🔴 Esta en CERO en las tres empresas. La contadora NO la usa: manda
        # esos minutos a la columna de 1,50. El modulo si la calcula, y por eso
        # paga de mas los mismos minutos.
        filas = leer_excel(cands[0])
        exc = sum(f["excedente"] for f in filas)
        print(f"\n[{empresa}] excedente 2.625 en el Excel: {exc:.2f}"
              + ("   ← la contadora NUNCA usa esa columna" if not exc else ""))

        # ── 2. Horas en cuartos: toda hora del Excel es multiplo de 0,25 ───
        horas, sueltas = [], []
        for f in filas:
            for campo in ("ex125", "ex150", "domingos", "feriados"):
                fx = f["formulas"].get(campo)
                if isinstance(fx, str):
                    for h in re.findall(r"[\(=+]\s*([\d.]+)\s*\*", fx):
                        v = float(h)
                        if v:
                            (horas if abs(v * 4 - round(v * 4)) < 1e-9 else sueltas).append(v)
        print(f"[{empresa}] horas tecleadas en las formulas: {len(horas)} en cuartos exactos, "
              f"{len(sueltas)} fuera de cuarto {sorted(set(sueltas)) if sueltas else ''}")

        # ── 3. «Desc. por fiesta judia»: la deuda de horas por los cierres ──
        # El local cierra en las fiestas y la gente debe esas horas. El Excel
        # las cobra aplicandoles la hora extra en vez de pagarsela. El modulo no
        # tiene el concepto: paga la hora extra completa.
        deuda = 0.0
        for hoja in wf.sheetnames:
            sf, sv = wf[hoja], wv[hoja]
            for r in range(1, sf.max_row + 1):
                et = sin_tildes(sf.cell(r, 1).value or "")
                if "FIESTA JUDIA" in et and num(sv.cell(r, 2).value):
                    saldo = None
                    for r2 in range(r, min(r + 10, sf.max_row + 1)):
                        if "PENDIENTE" in sin_tildes(sf.cell(r2, 1).value or "") or \
                           "HORAS A PAGAR" in sin_tildes(sf.cell(r2, 1).value or ""):
                            saldo = num(sv.cell(r2, 2).value)
                    emp = sv.cell(7, 2).value
                    d0 = num(sv.cell(r, 2).value)
                    deuda += saldo if saldo is not None else d0
                    print(f"[{empresa}] fiesta judia · hoja {hoja:5} {str(emp)[:24]:24} "
                          f"deuda {d0:7.2f} · aplicado {d0 - (saldo or 0):6.2f} · saldo {saldo or 0:7.2f}")
        if deuda:
            print(f"[{empresa}] saldo vivo de fiesta judia al cierre: {deuda:.2f}")

        # ── 4. Formulas de deducciones que se saltan una columna ───────────
        # Latente: hoy esas celdas estan vacias y no cambia un centavo, pero la
        # fila que las tenga llenas no descontaria.
        for f in filas:
            fx = f["formulas"].get("totded")
            if isinstance(fx, str):
                r = f["fila"]
                refs = set(re.findall(r"([A-Z]+)(\d+)", fx))
                malas = [c + n for c, n in refs if n != str(r)]
                faltan = [c for c in "MNOPQR" if (c, str(r)) not in refs]
                if malas or faltan:
                    print(f"[{empresa}] ⚠️  {f['nombre'][:24]:24} 'Total deducciones' = {fx}"
                          f"   {'apunta a otra fila: ' + ','.join(malas) if malas else ''}"
                          f"   {'omite ' + ','.join(faltan) if faltan else ''}")


if __name__ == "__main__":
    main()
