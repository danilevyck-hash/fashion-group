"""Tercer lector de los .xlsx: openpyxl. Otro programa, otro lenguaje.

🔴 POR QUE HACE FALTA. `xlsx-js-style` es el que ESCRIBE los archivos, asi que
leerlos con el solo prueba que sabe releerse a si mismo; el XML crudo prueba que
los bytes estan, pero lo interpreta el mismo que los escribio. `openpyxl` es el
parser con el que en este repo ya se abrio el Excel REAL de la contadora, y es
el que lee `freeze_panes` y `auto_filter` como los lee Excel: si aca no dice
`A2`, la fila de encabezados no se queda fija en la pantalla de nadie.

Solo lectura.

    npx tsx scripts/_verif-excel-panel-fijo.ts
    python3 scripts/_verif-excel-panel-fijo-openpyxl.py /tmp/excel-panel-fijo
"""
import glob
import os
import sys

import openpyxl

carpeta = sys.argv[1] if len(sys.argv) > 1 else "/tmp/excel-panel-fijo"
archivos = sorted(glob.glob(os.path.join(carpeta, "*.xlsx")))

if not archivos:
    print("🔴 no hay .xlsx en", carpeta, "— el verificador no probo nada")
    sys.exit(1)

AVISO = ("del 25 jul al 10 ago 2026 · NO es una quincena: "
         "sueldo base al 110.4 % y SIN los montos escritos a mano")

fallos = []
oks = 0

for ruta in archivos:
    nombre = os.path.basename(ruta)
    try:
        wb = openpyxl.load_workbook(ruta)
    except Exception as e:  # noqa: BLE001 — si no abre aca, no sirve
        fallos.append(f"{nombre}: openpyxl NO pudo abrirlo — {e}")
        continue

    for hoja in wb.sheetnames:
        ws = wb[hoja]
        etiqueta = f"{nombre} · {hoja}"

        ref = ws.auto_filter.ref
        # 🔑 UNA HOJA CON LAYOUT PROPIO NO LLEVA FILTRO NI PANEL, y esta bien: la
        # plantilla del banco B2B («DASHBOARD DE BUSQUEDA») tiene su rotulo
        # naranja en B1 y los codigos en B2:B201, asi que congelarle la fila 1
        # rompe el formato que el portal de PVH espera. Lo que si se exige es
        # COHERENCIA: sin filtro tampoco puede haber panel.
        if not ref:
            if ws.freeze_panes is not None:
                fallos.append(f"{etiqueta}: se congelo una hoja que no tiene encabezados")
            else:
                print(f"✅ {etiqueta}: layout propio — sin filtro y sin panel, como debe ser")
                oks += 1
            continue
        if ws.freeze_panes != "A2":
            fallos.append(f"{etiqueta}: freeze_panes = {ws.freeze_panes!r}, deberia ser 'A2'")
            continue
        if not ref.startswith("A1:"):
            fallos.append(f"{etiqueta}: auto_filter = {ref!r}, deberia arrancar en A1")
            continue
        a1 = ws["A1"].value
        if a1 is None or str(a1).strip() == "":
            fallos.append(f"{etiqueta}: A1 esta vacia — hay algo antes de los encabezados")
            continue

        print(f"✅ {etiqueta}: A1={a1!r} · filtro {ref} · freeze_panes {ws.freeze_panes}")
        oks += 1

# ── el aviso de la planilla, leido por el tercer parser ─────────────────────
planilla = os.path.join(carpeta, "planilla-rango-libre.xlsx")
if os.path.exists(planilla):
    ws = openpyxl.load_workbook(planilla)["Planilla"]
    filas = [c.row for f in ws.iter_rows() for c in f if c.value == AVISO]
    fin_filtro = int(ws.auto_filter.ref.split(":")[1].lstrip("ABCDEFGHIJKLMNOPQRSTUVWXYZ"))
    if len(filas) != 1:
        fallos.append(f"el aviso de la planilla aparece {len(filas)} veces (deberia aparecer 1)")
    elif filas[0] <= fin_filtro:
        fallos.append(f"el aviso esta DENTRO del filtro (fila {filas[0]} <= {fin_filtro})")
    else:
        print(f"✅ el aviso de la planilla en la fila {filas[0]}, FUERA del filtro (termina en {fin_filtro})")
        oks += 1
else:
    fallos.append("falta planilla-rango-libre.xlsx — el aviso no se verifico")

print("")
for f in fallos:
    print("🔴", f)
if fallos:
    print(f"\n🔴 {len(fallos)} hallazgo(s) con openpyxl {openpyxl.__version__}")
    sys.exit(1)
print(f"🟢 {oks} verificaciones OK con openpyxl {openpyxl.__version__}, sobre {len(archivos)} archivos")
