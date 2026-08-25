"""Segundo parser del Excel de la Planilla: openpyxl.

🔴 POR QUE DOS PARSERS. `xlsx-js-style` es el que ESCRIBE el archivo, asi que
leerlo con el solo prueba que sabe releerse a si mismo. `openpyxl` es otro
programa, en otro lenguaje, y con el se abrio el Excel REAL de la contadora:
si el nuestro no abre ahi, no sirve.

Lee el archivo que bajo `_verif-planilla-excel-ausencia.mjs` desde la app real.
Solo lectura.

    python3 scripts/_verif-planilla-excel-openpyxl.py /tmp/excel-ausencia/planilla-boston.xlsx
"""
import sys
import openpyxl

ruta = sys.argv[1] if len(sys.argv) > 1 else "/tmp/excel-ausencia/planilla-boston.xlsx"
problemas = []

# data_only=False es como se leyo el Excel de la contadora: muestra las formulas.
# El nuestro no tiene formulas -escribe valores- y eso tambien hay que verlo.
wb = openpyxl.load_workbook(ruta, data_only=False)
print("hojas:", wb.sheetnames)
if len(wb.sheetnames) < 3:
    problemas.append("el archivo trae menos de 3 hojas")

def filas(ws):
    return [[c.value for c in f] for f in ws.iter_rows()]

# ── Hoja 1: el cuadro ────────────────────────────────────────────────────────
cuadro = filas(wb["Planilla"])
enc = next((f for f in cuadro if "Total bruto" in [str(x) for x in f]), None)
if enc is None:
    problemas.append("no encontre el encabezado del cuadro")
else:
    col = {str(v): i for i, v in enumerate(enc) if v is not None}
    NECESARIAS = ["Salario quincenal", "Horas extra 1.25", "Ausencias", "Tardanzas",
                  "Horas extra 1.50", "Excedente 2.625", "Domingos", "Feriados", "Total bruto"]
    faltan = [n for n in NECESARIAS if n not in col]
    if faltan:
        problemas.append("faltan columnas en el cuadro: " + " · ".join(faltan))
    else:
        def n(f, k):
            v = f[col[k]]
            return float(v) if isinstance(v, (int, float)) else 0.0
        revisadas = 0
        for f in cuadro[cuadro.index(enc) + 1:]:
            bruto = f[col["Total bruto"]]
            if not isinstance(bruto, (int, float)) or bruto == 0:
                continue
            esperado = round(
                n(f, "Salario quincenal") + n(f, "Horas extra 1.25") + n(f, "Horas extra 1.50")
                + n(f, "Excedente 2.625") + n(f, "Domingos") + n(f, "Feriados")
                - n(f, "Ausencias") - n(f, "Tardanzas"), 2)
            revisadas += 1
            if abs(esperado - float(bruto)) > 0.005:
                problemas.append(f"{f[0]}: el bruto ({bruto}) no cuadra con sus columnas ({esperado})")
        print("filas del cuadro con bruto verificado a mano:", revisadas)
        if revisadas < 5:
            problemas.append("se verificaron menos de 5 filas: el Excel salio casi vacio")

# ── Hoja 2: las horas, con las columnas nuevas ───────────────────────────────
horas = filas(wb["Horas"])
enc_h = next((f for f in horas if "Tarde >30 min (min)" in [str(x) for x in f]), None)
if enc_h is None:
    problemas.append("la hoja de horas no tiene «Tarde >30 min (min)»")
else:
    col_h = {str(v): i for i, v in enumerate(enc_h) if v is not None}
    if "Tarde >30 min (días)" not in col_h:
        problemas.append("falta la columna «Tarde >30 min (días)»")
    else:
        con_grave = 0
        for f in horas[horas.index(enc_h) + 1:]:
            d = f[col_h["Tarde >30 min (días)"]]
            m = f[col_h["Tarde >30 min (min)"]]
            if isinstance(d, (int, float)) and d > 0:
                con_grave += 1
                if not isinstance(m, (int, float)) or m <= 30:
                    problemas.append(f"{f[0]}: dice {d} dia(s) de mas de 30 min pero los minutos son {m}")
        print("personas con dias de mas de 30 minutos tarde:", con_grave)
        if con_grave == 0:
            problemas.append("ninguna fila trae dias de mas de 30 minutos tarde: el archivo no prueba nada")

# ── Hoja 3: la explicacion tiene que estar EN EL PAPEL ──────────────────────
como = " ".join(str(c) for f in filas(wb["Cómo se calcula"]) for c in f if c is not None)
for frase in ["Llegar más de 30 minutos tarde", "SE DESCUENTAN LOS MINUTOS",
              "el total bruto y el neto son los mismos"]:
    if frase not in como:
        problemas.append(f"la hoja «Cómo se calcula» no dice: {frase}")

if problemas:
    print("\n🔴 " + "\n🔴 ".join(problemas))
    sys.exit(1)
print("\n🟢 openpyxl: las 3 hojas abren, las columnas nuevas están, el bruto cuadra"
      " columna por columna y la explicación viaja en el papel.")
